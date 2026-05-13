#!/usr/bin/env python3
"""Build comparison_5-6-26.xlsx from PDF data and EXCEL_ACTIONS dict."""

import glob
import re
import subprocess
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pypdf

from boss_actions import EXCEL_ACTIONS

# ── PDF extraction ─────────────────────────────────────────────────────────────

SCORE_RE    = re.compile(r'^(\d{2,3})\n/ 100', re.MULTILINE)
# New format: action at line start, before score block
ACT_RE_NEW  = re.compile(r'^(STAGED BUY|BUY NOW|BUY|HOLD|WAIT|SELL)', re.MULTILINE)
# Old format: "ACTION Thesis: <thesis>→ <action>"
ACT_RE_OLD  = re.compile(
    r'ACTION Thesis:[^→\n]*→\s*(STAGED BUY|BUY NOW|BUY|HOLD|WAIT|SELL)', re.MULTILINE
)


def _extract_pdf(path: str) -> dict:
    p = pypdf.PdfReader(path)
    text = "\n".join(page.extract_text() or "" for page in p.pages[:4])

    m_score = SCORE_RE.search(text)
    score = int(m_score.group(1)) if m_score else None

    # Old format detection: header line contains "ACTION Thesis:"
    if "ACTION Thesis:" in text:
        m_act = ACT_RE_OLD.search(text)
    else:
        # New format: action appears before the score block
        # Restrict search to text before "/ 100" to avoid "BUY TRIGGER" section headers
        pre_score = text[: m_score.start()] if m_score else text[:500]
        m_act = ACT_RE_NEW.search(pre_score)

    return {
        "score":  score,
        "action": m_act.group(1) if m_act else None,
    }


def load_pdf_data() -> dict:
    """Return {ticker: {score, action}} preferring fixed batch, falling back to older batches."""
    data: dict[str, dict] = {}

    # Priority order: fixed batch, then b batch, then original batch
    for batch_glob, date_str in [
        ("reports/batch_2026-05-06_fixed/*.pdf", "2026-05-06"),
        ("reports/batch_2026-05-05b/*.pdf",      "2026-05-05"),
        ("reports/batch_2026-05-05/*.pdf",        "2026-05-05"),
    ]:
        for pdf_path in sorted(glob.glob(batch_glob)):
            fname  = Path(pdf_path).stem          # e.g. "AAPL_2026-05-06"
            ticker = fname.split("_")[0]
            if ticker in data:
                continue                          # already have a better source
            try:
                result = _extract_pdf(pdf_path)
                data[ticker] = result
            except Exception as e:
                print(f"  WARN: could not read {pdf_path}: {e}")

    return data


# ── Hard-coded fallbacks (tickers with no PDF) ────────────────────────────────
# From batch_2026-05-05b task output (evaluated but PDF timed out / not saved)

FALLBACK: dict[str, dict] = {
    # Evaluated, no PDF (blocked or save failure) — data from batch task output
    "NVDA": {"score": 79, "action": "BUY"},
    "CHTR": {"score": None, "action": None},
    "D":    {"score": None, "action": None},
    "EIX":  {"score": None, "action": None},
    "F":    {"score": None, "action": None},
    "TM":   {"score": None, "action": None},
    # Never evaluated
    "BABA": {"score": None, "action": None},
    "BRK.B":{"score": None, "action": None},
}


# ── Client recommendation logic ───────────────────────────────────────────────

def client_recommendation(se_action: str | None, excel_action: str) -> tuple[str, str]:
    """Returns (recommendation, reasoning)."""
    if se_action is None:
        return ("No data", "Evaluation unavailable — check manually.")

    # Normalise display: WAIT → HOLD for StockEval display
    # But use raw action for logic
    act = se_action.upper().strip()

    if act == "SELL":
        return (
            "Sell / Reduce",
            "Model rates as SELL — consider trimming or exiting the position.",
        )
    if act in ("BUY", "BUY NOW"):
        return (
            "Add to Position",
            "Strong entry signal — risk/return supports increasing allocation now.",
        )
    if act == "STAGED BUY":
        return (
            "Add Gradually",
            "Long-term buy case confirmed; build in tranches rather than all at once.",
        )
    if act == "HOLD":
        if excel_action.upper() == "SELL":
            return (
                "Hold — consider reducing",
                "Model rates fair value but Excel rates SELL — review for potential trim.",
            )
        return (
            "Hold",
            "Fair value — maintain current allocation; no strong catalyst to add or reduce.",
        )
    if act == "WAIT":
        # Distinguish bullish vs cautious WAIT
        if excel_action.upper() in ("BUY",):
            return (
                "Hold — don't add yet",
                "Excel rates BUY but model says wait for a better entry point; keep what you have.",
            )
        if excel_action.upper() == "SELL":
            return (
                "Hold — consider reducing",
                "Model cautions against entry; Excel rates SELL — review for potential trim.",
            )
        return (
            "Hold — don't add yet",
            "Risk/return not compelling right now; hold existing position and revisit.",
        )

    return ("Hold", "Review manually.")


# ── Excel styling helpers ─────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold: bool = False, color: str = "000000", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border() -> Border:
    side = Side(style="thin", color="D0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)

ACTION_COLORS = {
    "BUY":        ("1E6B1E", "FFFFFF"),   # dark green / white text
    "BUY NOW":    ("1E6B1E", "FFFFFF"),
    "STAGED BUY": ("2E7D32", "FFFFFF"),   # medium green
    "HOLD":       ("5D4037", "FFFFFF"),   # brown
    "WAIT":       ("616161", "FFFFFF"),   # grey — same display as HOLD
    "SELL":       ("B71C1C", "FFFFFF"),   # red
}

REC_COLORS = {
    "Sell / Reduce":         ("FFCDD2", "000000"),
    "Add to Position":       ("C8E6C9", "000000"),
    "Add Gradually":         ("DCEDC8", "000000"),
    "Hold":                  ("FFF9C4", "000000"),
    "Hold — don't add yet":  ("FFF9C4", "000000"),
    "Hold — consider reducing": ("FFE0B2", "000000"),
    "No data":               ("EEEEEE", "888888"),
}

MATCH_COLORS = {
    True:  ("E8F5E9", "2E7D32"),   # light green
    False: ("FBE9E7", "B71C1C"),   # light red
}


def action_display(action: str | None) -> str:
    """WAIT is shown as HOLD on the sheet."""
    if action is None:
        return "—"
    if action.upper() == "WAIT":
        return "HOLD"
    return action.upper()


def actions_match(excel: str, se: str | None) -> bool | None:
    if se is None:
        return None
    e = excel.upper().strip()
    s = se.upper().strip()
    # WAIT counts as HOLD for comparison purposes
    if s == "WAIT":
        s = "HOLD"
    if e == "BUY NOW":
        e = "BUY"
    return e == s


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    out_path = Path.home() / "Desktop" / "comparison_5-6-26.xlsx"

    print("Loading PDF data …")
    pdf_data = load_pdf_data()

    # Merge fallbacks
    for ticker, fallback in FALLBACK.items():
        if ticker not in pdf_data:
            pdf_data[ticker] = fallback

    # All tickers: union of EXCEL_ACTIONS and any extras evaluated
    all_tickers = sorted(EXCEL_ACTIONS.keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # ── Header row
    headers = [
        "Ticker", "Score", "Excel Action", "StockEval Action",
        "Match?", "Client Recommendation", "Reasoning",
    ]
    header_fill = _fill("1A3A6B")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
    ws.row_dimensions[1].height = 28

    # ── Data rows
    for row_idx, ticker in enumerate(all_tickers, 2):
        pdf   = pdf_data.get(ticker, {"score": None, "action": None})
        score = pdf.get("score")
        se_action_raw = pdf.get("action")
        excel_action  = EXCEL_ACTIONS.get(ticker, "—")
        se_display    = action_display(se_action_raw)
        match         = actions_match(excel_action, se_action_raw)
        rec, reasoning = client_recommendation(se_action_raw, excel_action)

        values = [ticker, score, excel_action, se_display, "☑" if match else "☐", rec, reasoning]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = _border()
            cell.alignment = Alignment(
                horizontal="center" if col != 7 else "left",
                vertical="center",
                wrap_text=(col == 7),
            )
            cell.font = _font()

            # Column-specific colouring
            if col == 3:   # Excel Action
                bg, fg = ACTION_COLORS.get(excel_action.upper(), ("FFFFFF", "000000"))
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=fg)
            elif col == 4:  # StockEval Action
                raw = (se_action_raw or "").upper()
                # WAIT displayed as HOLD but same colour bucket
                lookup = raw if raw != "WAIT" else "HOLD"
                bg, fg = ACTION_COLORS.get(lookup, ("FFFFFF", "000000"))
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=fg)
            elif col == 5:  # Match?
                if match is None:
                    cell.fill = _fill("EEEEEE")
                    cell.font = _font(color="888888")
                    cell.value = "—"
                else:
                    bg, fg = MATCH_COLORS[match]
                    cell.fill = _fill(bg)
                    cell.font = _font(bold=True, color=fg)
            elif col == 6:  # Client Recommendation
                bg, fg = REC_COLORS.get(rec, ("FFFFFF", "000000"))
                cell.fill = _fill(bg)

        ws.row_dimensions[row_idx].height = 52

    # ── Column widths
    widths = [10, 8, 14, 16, 9, 22, 55]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    wb.save(str(out_path))
    print(f"Saved: {out_path}")

    # ── Verify
    print("\nVerification:")
    wb2 = openpyxl.load_workbook(str(out_path))
    ws2 = wb2.active
    rows = list(ws2.iter_rows(min_row=2, values_only=True))
    print(f"  Rows: {len(rows)} (expected {len(all_tickers)})")

    with_score  = sum(1 for r in rows if r[1] is not None)
    with_action = sum(1 for r in rows if r[3] and r[3] != "—")
    waits       = sum(1 for r in rows if (r[3] or "").upper() == "WAIT")
    matches     = sum(1 for r in rows if r[4] == "☑")
    mismatches  = sum(1 for r in rows if r[4] == "☐")
    no_data     = sum(1 for r in rows if r[4] == "—")

    print(f"  Scores filled:  {with_score}/{len(rows)}")
    print(f"  Actions filled: {with_action}/{len(rows)}")
    print(f"  WAIT visible:   {waits}  (should be 0 — all displayed as HOLD)")
    print(f"  Match ☑: {matches}  Mismatch ☐: {mismatches}  No data —: {no_data}")

    print("\nSample rows:")
    for r in rows[:5]:
        print(f"  {r[:6]}")

    subprocess.run(["open", str(out_path)], check=False)


if __name__ == "__main__":
    main()
